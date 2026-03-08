import openpyxl
import os
import shutil

class ExcelInjector:
    def __init__(self, excel_path: str = "/app/data/SSTOK_LISTESI.xlsx", output_name: str = "SSTOK_LISTESI_GUNCEL.xlsx"):
        self.excel_path = excel_path
        
        # Yeni dosya yolunu belirle (GUNCEL olan)
        dir_name = os.path.dirname(excel_path)
        self.output_path = os.path.join(dir_name, output_name)
        
        # Orijinal dosyayı formülleriyle birlikte güvenle kopyala
        if os.path.exists(self.excel_path):
             shutil.copy2(self.excel_path, self.output_path)
             
        self.wb = None
        self.ws = None

    def _load_workbook(self):
         # data_only=False formüllerin uçmaması içindir.
         if not self.wb:
             self.wb = openpyxl.load_workbook(self.output_path, data_only=False)
             self.ws = self.wb.active

    def inject_prices(self, updates: list):
        """
        updates listesi şu formattadır:
        [
            {"excel_row_index": 1205, "new_price": 54.20, "price_column": "Q"},
            {"excel_row_index": 140, "new_price": 10.50, "price_column": "Q"}
        ]
        
        Not: pandas row_index 0'dan başlar. Excel'de 1. satır başlıktır. Data 2. satırdan başlar.
        Yani excel_row_index = 0 demek Excel'de 2. satıra denk gelir.
        Dolayısıyla: Openpyxl Row = excel_row_index + 2 olur.
        """
        if not updates:
            print("Enjekte edilecek güncelleme bulunamadı.")
            return

        self._load_workbook()
        
        success_count = 0
        for update in updates:
            row_idx = update["excel_row_index"] + 2
            col_letter = update.get("price_column", "Q") # İskontosuz Perakende sütunu olabilir vs.
            new_price = update["new_price"]
            
            target_cell = f"{col_letter}{row_idx}"
            self.ws[target_cell] = new_price
            success_count += 1
            
        self.wb.save(self.output_path)
        print(f"✅ Excel Enjeksiyonu Başarılı: Toplam {success_count} hücre güncellendi ve formüller korundu.")
        print(f"📂 Dosya şuraya kaydedildi: {self.output_path}")

